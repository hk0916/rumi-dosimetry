#!/usr/bin/env python3
"""Dosimetry 웹 애플리케이션 개발 플랜 - Excel 문서 생성"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 공통 스타일 ──
HEADER_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUB_HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
SUB_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CATEGORY_FONT = Font(name="Arial", size=11, bold=True)
CATEGORY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
NORMAL_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=16, bold=True, color="2F5496")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols, font=HEADER_FONT, fill=HEADER_FILL):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_data_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER


def style_category_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = CATEGORY_FONT
        cell.fill = CATEGORY_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER


# ═══════════════════════════════════════════
# Sheet 1: 요구사항 정의
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. 요구사항 정의"
ws1.sheet_properties.tabColor = "2F5496"

ws1.merge_cells("A1:D1")
ws1["A1"] = "Dosimetry 웹 애플리케이션 - 요구사항 정의서"
ws1["A1"].font = TITLE_FONT
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 40

headers = ["카테고리", "기능", "설명", "우선순위"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, 4)

data = [
    # 인증/사용자 관리
    ("인증/사용자 관리", None, None, None),
    ("", "로그인/로그아웃", "ID/PW 기반 인증, 세션 관리", "P1"),
    ("", "사용자 CRUD", "사용자 추가/수정/삭제, 역할 관리 (관리자/최고관리자)", "P5"),
    ("", "내 프로필", "사용자 이름, 권한, 그룹 확인 및 비밀번호 변경", "P5"),
    ("", "워크스페이스 설정", "워크스페이스 단위 관리", "P5"),
    # 디바이스 관리
    ("디바이스 관리", None, None, None),
    ("", "Dosimeter 목록", "Device Name, Status, Device Type, MAC Address, Battery, RSSI, Voltage, Uptime 테이블 (페이지네이션)", "P1"),
    ("", "Dosimeter CRUD", "디바이스 추가/상세보기(모달)/수정(Setting)/삭제", "P1"),
    ("", "Gateway 목록", "Gateway 디바이스 목록 (Device Name, Status, Device Type, MAC Address, Server IP, Uptime)", "P1"),
    ("", "Gateway 상세/설정", "[Device 탭] MAC Address, Device Type, Server IP, Status, Uptime, Server URL, IPv4 모드, IP/Subnet/GW/DNS, BLE RSSI Threshold, Device FW 버전, BLE FW 버전, LED 상태\n[Setting 탭] 네트워크 수정, Interface, LED On/Off, BLE RSSI 수정, Device FW Upload, BLE FW Upload", "P1"),
    # Data Monitoring
    ("Data Monitoring", None, None, None),
    ("", "디바이스 선택", "드롭다운으로 Dosimeter 선택", "P2"),
    ("", "실시간 차트", "선택된 디바이스의 Voltage(mV) vs Time 실시간 스트리밍 차트", "P2"),
    ("", "Start/Stop", "모니터링 시작/중지 제어", "P2"),
    # Calibration
    ("Calibration (교정)", None, None, None),
    ("", "데이터 조회", "디바이스 선택 + 날짜/시간 범위 지정 → Import Data", "P3"),
    ("", "스무딩 필터", "Median, Arithmetic Mean, Geometric Mean, Least Square, Envelope, Bezier (6종)", "P3"),
    ("", "Window Size / Baseline", "필터 윈도우 사이즈 및 베이스라인 값 설정", "P3"),
    ("", "차트 렌더링", "Original(파랑) + Filtered(초록, 원본-baseline) 차트 동시 표시", "P3"),
    ("", "누적선량 계산", "Start~End Time 구간의 V*s 적분값(Cumulative Dose) 계산", "P3"),
    ("", "CF Factor 산출", "Delivered Dose(cGy) 입력 → CF Factor(mV/cGy) 자동 계산 및 저장", "P3"),
    # Manage Calibration
    ("Manage Calibration", None, None, None),
    ("", "CF Factor 목록", "저장된 Calibration Factor 목록 조회", "P4"),
    ("", "상세 조회", "누적 선량, 필터 세팅, 차트 재확인", "P4"),
    ("", "수정/삭제", "저장된 교정 데이터 관리", "P4"),
    # Data Analysis
    ("Data Analysis", None, None, None),
    ("", "방사선 종류 선택", "Photon, Electron & Muon, Proton, Neutron, Alpha Particle/Fission Fragment/Heavy Ion", "P4"),
    ("", "대상 장기 선택", "Breast, Colon, Stomach, Lung, Gonads, Bladder, Liver, Esophagus, Thyroid, Bone Surface, Brain, Salivary Glands, Skin, Residual Tissues", "P4"),
    ("", "범위 설정", "Full Range 또는 Sub Range(사용자 지정 시간 구간)", "P4"),
    ("", "선량 계산", "Cumulative Dose(누적선량), Absorbed Dose(흡수선량), Equivalent Dose(등가선량), Effective Dose(유효선량)", "P4"),
    ("", "Export", ".csv 파일 다운로드", "P4"),
]

row = 4
for d in data:
    if d[1] is None:  # category row
        ws1.cell(row=row, column=1, value=d[0])
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        style_category_row(ws1, row, 4)
    else:
        for c, v in enumerate(d, 1):
            ws1.cell(row=row, column=c, value=v)
        style_data_row(ws1, row, 4)
    row += 1

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 65
ws1.column_dimensions["D"].width = 12

# ═══════════════════════════════════════════
# Sheet 2: 화면 설계
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("2. 화면 설계")
ws2.sheet_properties.tabColor = "548235"

ws2.merge_cells("A1:E1")
ws2["A1"] = "Dosimetry 웹 애플리케이션 - 화면 설계 (UI/UX)"
ws2["A1"].font = TITLE_FONT
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 40

headers2 = ["화면명", "URL 경로", "주요 컴포넌트", "사용자 액션", "연관 API"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 5)

screens = [
    ("Login", "/login", "로고, ID 입력, PW 입력, 로그인 버튼", "ID/PW 입력 → 로그인", "POST /api/auth/login"),
    ("Device - Dosimeter", "/dosimetry/device\n(Dosimeter 탭)", "테이블(Device Name, Status, Type, MAC, Battery, RSSI, Voltage, Uptime)\n페이지네이션, Add Device 버튼", "행 클릭 → 상세 모달\nAdd Device → 추가 모달", "GET /api/devices\nPOST /api/devices\nPUT /api/devices/:id\nDELETE /api/devices/:id"),
    ("Device - Gateway", "/dosimetry/device\n(Gateway 탭)", "테이블(Device Name, Status, Type, MAC, Server IP, Uptime)\n[Device 탭] MAC, Type, ServerIP, Status, Uptime, Server URL, IPv4(Manual/Auto), IP/Subnet/GW/DNS, BLE RSSI Threshold, Device FW, BLE FW, LED\n[Setting 탭] 네트워크 설정, Interface, LED On/Off, BLE RSSI Threshold 수정, Device FW Upload, BLE FW Upload", "행 클릭 → Device 탭(조회)\nSetting 탭 → 설정 변경\nFW Upload → 펌웨어 업로드", "GET /api/gateways\nPUT /api/gateways/:id/settings\nPOST /api/gateways/:id/firmware"),
    ("User 설정", "/dosimetry/device\n(Setting 모달)", "내 프로필(이름, 권한, 그룹)\n사용자 목록(사용자 탭/그룹 탭)\n계정 및 보안(비밀번호 변경)", "프로필 변경\n사용자 추가/삭제\n비밀번호 변경", "GET /api/users\nPOST /api/users\nPUT /api/users/:id"),
    ("Data Monitoring", "/dosimetry/data-monitoring", "디바이스 드롭다운\nStart/Stop 버튼\nVoltage vs Time 실시간 차트", "디바이스 선택 → Start → 실시간 차트 표시\nStop → 중지", "WS /ws/monitoring/:deviceId"),
    ("Calibration", "/dosimetry/calibration", "Header: Filter 드롭다운, Window Size, Baseline, Apply\n디바이스 선택, 날짜/시간 범위, Import Data\n차트: Original(파랑) + Filtered(초록)\nFooter: Start/End Time, Cumulative Dose, Delivered Dose, Result", "1. 디바이스/시간 선택 → Import Data\n2. 필터/윈도우/베이스라인 → Apply\n3. Start/End Time → 누적선량 계산\n4. Delivered Dose 입력 → CF Factor → Save", "GET /api/sensor-data\nPOST /api/calibrations/calculate\nPOST /api/calibrations"),
    ("Manage Calibration", "/dosimetry/manage-calibration", "CF Factor 목록 테이블\n선택 시 차트 + 세팅값 표시\nFooter 계산 영역", "CF Factor 선택 → 상세 확인\n수정/삭제", "GET /api/calibrations\nGET /api/calibrations/:id\nDELETE /api/calibrations/:id"),
    ("Data Analysis", "/dosimetry/data-analysis", "Header: Filter, Window Size, Baseline, Apply\n방사선 종류 드롭다운, 대상 장기 드롭다운\n차트 표시\nFooter: Full Range/Sub Range, Calculate, Report, Export", "1. 디바이스/시간 선택 → Import Data\n2. 방사선 종류 + 장기 선택\n3. Calculate → 4종 선량 결과\n4. Export → CSV 다운로드", "POST /api/analysis/calculate\nGET /api/analysis/export"),
]

for i, s in enumerate(screens):
    r = 4 + i
    for c, v in enumerate(s, 1):
        ws2.cell(row=r, column=c, value=v)
    style_data_row(ws2, r, 5)
    ws2.row_dimensions[r].height = 80

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 25
ws2.column_dimensions["C"].width = 45
ws2.column_dimensions["D"].width = 40
ws2.column_dimensions["E"].width = 30

# ═══════════════════════════════════════════
# Sheet 3: DB 설계
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("3. DB 설계")
ws3.sheet_properties.tabColor = "BF8F00"

ws3.merge_cells("A1:F1")
ws3["A1"] = "Dosimetry 웹 애플리케이션 - DB 설계"
ws3["A1"].font = TITLE_FONT
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 40

tables = [
    ("users (사용자)", [
        ("id", "BIGINT", "PK, AUTO_INCREMENT", "사용자 고유 ID", ""),
        ("username", "VARCHAR(50)", "UNIQUE, NOT NULL", "로그인 ID", ""),
        ("password_hash", "VARCHAR(255)", "NOT NULL", "암호화된 비밀번호", ""),
        ("name", "VARCHAR(100)", "", "사용자 이름", ""),
        ("role", "ENUM", "", "권한: admin / super_admin", ""),
        ("account_type", "VARCHAR(50)", "", "계정 타입 (Local Account 등)", ""),
        ("group_name", "VARCHAR(100)", "", "소속 그룹", ""),
        ("workspace_id", "BIGINT", "FK", "워크스페이스 ID", "→ workspaces.id"),
        ("created_at", "DATETIME", "", "생성일시", ""),
        ("updated_at", "DATETIME", "", "수정일시", ""),
    ]),
    ("workspaces (워크스페이스)", [
        ("id", "BIGINT", "PK, AUTO_INCREMENT", "워크스페이스 고유 ID", ""),
        ("name", "VARCHAR(100)", "", "워크스페이스 이름", ""),
        ("created_at", "DATETIME", "", "생성일시", ""),
    ]),
    ("devices (Dosimeter 디바이스)", [
        ("id", "BIGINT", "PK, AUTO_INCREMENT", "디바이스 고유 ID", ""),
        ("device_name", "VARCHAR(100)", "NOT NULL", "디바이스 이름", ""),
        ("device_type", "VARCHAR(50)", "", "디바이스 타입 (Skin Dosimeter 등)", ""),
        ("mac_address", "VARCHAR(17)", "UNIQUE", "MAC 주소", ""),
        ("status", "ENUM", "DEFAULT 'offline'", "상태: online / offline", ""),
        ("battery", "INT", "", "배터리 (%)", ""),
        ("rssi", "INT", "", "신호 세기 (dBm)", ""),
        ("voltage", "DECIMAL(10,4)", "", "전압 (mV)", ""),
        ("uptime", "DATETIME", "", "최근 online 전환 시간", ""),
        ("workspace_id", "BIGINT", "FK", "워크스페이스 ID", "→ workspaces.id"),
        ("created_at", "DATETIME", "", "생성일시", ""),
        ("updated_at", "DATETIME", "", "수정일시", ""),
    ]),
    ("gateways (Gateway 디바이스)", [
        ("id", "BIGINT", "PK, AUTO_INCREMENT", "Gateway 고유 ID", ""),
        ("device_name", "VARCHAR(100)", "NOT NULL", "디바이스 이름", ""),
        ("device_type", "VARCHAR(50)", "", "타입 (Twin Tracker BLE 등)", ""),
        ("mac_address", "VARCHAR(17)", "UNIQUE", "MAC 주소", ""),
        ("status", "ENUM", "", "상태: online / offline", ""),
        ("server_ip", "VARCHAR(45)", "", "서버 IP", ""),
        ("server_url", "VARCHAR(255)", "", "서버 URL", ""),
        ("ipv4_mode", "ENUM", "DEFAULT 'manual'", "IPv4 모드: manual / auto", ""),
        ("ip_address", "VARCHAR(45)", "", "IP 주소", ""),
        ("subnet_mask", "VARCHAR(45)", "", "서브넷 마스크", ""),
        ("gateway_ip", "VARCHAR(45)", "", "게이트웨이 IP", ""),
        ("dns_main", "VARCHAR(45)", "", "주 DNS", ""),
        ("dns_sub", "VARCHAR(45)", "", "보조 DNS", ""),
        ("interface_type", "VARCHAR(50)", "", "인터페이스 타입", ""),
        ("led_enabled", "BOOLEAN", "DEFAULT true", "LED On/Off 설정", ""),
        ("ble_rssi_threshold", "INT", "", "BLE RSSI 임계값", ""),
        ("device_fw_version", "VARCHAR(50)", "", "Device 펌웨어 버전", ""),
        ("ble_fw_version", "VARCHAR(50)", "", "BLE 펌웨어 버전", ""),
        ("uptime", "DATETIME", "", "최근 online 전환 시간", ""),
        ("workspace_id", "BIGINT", "FK", "워크스페이스 ID", "→ workspaces.id"),
        ("created_at", "DATETIME", "", "생성일시", ""),
        ("updated_at", "DATETIME", "", "수정일시", ""),
    ]),
    ("sensor_data (센서 원시 데이터)", [
        ("id", "BIGINT", "PK, AUTO_INCREMENT", "데이터 고유 ID", ""),
        ("device_id", "BIGINT", "FK, INDEX", "디바이스 ID", "→ devices.id"),
        ("timestamp", "DATETIME(3)", "NOT NULL, INDEX", "측정 시간 (ms 정밀도)", ""),
        ("voltage", "DECIMAL(10,4)", "", "측정 전압 (mV)", ""),
        ("created_at", "DATETIME", "", "생성일시", ""),
    ]),
    ("calibrations (Calibration 결과)", [
        ("id", "BIGINT", "PK, AUTO_INCREMENT", "교정 고유 ID", ""),
        ("device_id", "BIGINT", "FK", "디바이스 ID", "→ devices.id"),
        ("user_id", "BIGINT", "FK", "수행 사용자 ID", "→ users.id"),
        ("date", "DATE", "", "교정 날짜", ""),
        ("filter_type", "ENUM", "", "필터: median/arithmetic_mean/geometric_mean/least_square/envelope/bezier", ""),
        ("window_size", "INT", "", "필터 윈도우 사이즈", ""),
        ("baseline", "DECIMAL(10,4)", "", "베이스라인 값", ""),
        ("start_time", "DATETIME(3)", "", "계산 시작 시간", ""),
        ("end_time", "DATETIME(3)", "", "계산 종료 시간", ""),
        ("cumulative_dose", "DECIMAL(15,4)", "", "누적선량 (V·s)", ""),
        ("delivered_dose", "DECIMAL(10,4)", "", "전달선량 (cGy)", ""),
        ("cf_factor", "DECIMAL(10,4)", "", "CF Factor (mV/cGy)", ""),
        ("cf_name", "VARCHAR(100)", "", "저장 이름", ""),
        ("created_at", "DATETIME", "", "생성일시", ""),
        ("updated_at", "DATETIME", "", "수정일시", ""),
    ]),
    ("analysis_results (Data Analysis 결과)", [
        ("id", "BIGINT", "PK, AUTO_INCREMENT", "분석 결과 고유 ID", ""),
        ("calibration_id", "BIGINT", "FK", "교정 ID", "→ calibrations.id"),
        ("user_id", "BIGINT", "FK", "수행 사용자 ID", "→ users.id"),
        ("radiation_source", "ENUM", "", "방사선 종류: photon/electron_muon/proton/neutron/alpha_heavy_ion", ""),
        ("target_organ", "VARCHAR(50)", "", "대상 장기", ""),
        ("range_type", "ENUM", "", "범위: full / sub", ""),
        ("sub_range_start", "DATETIME(3)", "", "서브 범위 시작 (sub인 경우)", ""),
        ("sub_range_end", "DATETIME(3)", "", "서브 범위 종료 (sub인 경우)", ""),
        ("cumulative_dose", "DECIMAL(15,6)", "", "누적선량", ""),
        ("absorbed_dose", "DECIMAL(15,6)", "", "흡수선량 (Gy)", ""),
        ("equivalent_dose", "DECIMAL(15,6)", "", "등가선량 (Sv)", ""),
        ("effective_dose", "DECIMAL(15,6)", "", "유효선량 (Sv)", ""),
        ("created_at", "DATETIME", "", "생성일시", ""),
    ]),
    ("radiation_weighting_factors (방사선 가중치)", [
        ("id", "BIGINT", "PK, AUTO_INCREMENT", "고유 ID", ""),
        ("radiation_source", "VARCHAR(50)", "", "방사선 종류", ""),
        ("weighting_factor", "DECIMAL(10,4)", "", "방사선 가중치 (wR)", ""),
    ]),
    ("tissue_weighting_factors (조직 가중치)", [
        ("id", "BIGINT", "PK, AUTO_INCREMENT", "고유 ID", ""),
        ("organ_name", "VARCHAR(50)", "", "장기 이름", ""),
        ("weighting_factor", "DECIMAL(10,6)", "", "조직 가중치 (wT)", ""),
    ]),
]

col_headers = ["컬럼명", "데이터 타입", "제약조건", "설명", "참조"]
row = 3
for tbl_name, cols in tables:
    ws3.cell(row=row, column=1, value=tbl_name)
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    style_category_row(ws3, row, 6)
    ws3.row_dimensions[row].height = 28
    row += 1

    for i, h in enumerate(col_headers, 1):
        ws3.cell(row=row, column=i, value=h)
    style_header_row(ws3, row, 5, SUB_HEADER_FONT, SUB_HEADER_FILL)
    row += 1

    for col_data in cols:
        for c, v in enumerate(col_data, 1):
            ws3.cell(row=row, column=c, value=v)
        style_data_row(ws3, row, 5)
        row += 1

    row += 1  # blank row between tables

ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 45
ws3.column_dimensions["E"].width = 20

# ═══════════════════════════════════════════
# Sheet 4: 시스템 설계
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("4. 시스템 설계")
ws4.sheet_properties.tabColor = "C00000"

ws4.merge_cells("A1:D1")
ws4["A1"] = "Dosimetry 웹 애플리케이션 - 시스템 설계"
ws4["A1"].font = TITLE_FONT
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 40

# 기술 스택
ws4.cell(row=3, column=1, value="기술 스택")
ws4.merge_cells("A3:D3")
style_category_row(ws4, 3, 4)

tech_headers = ["레이어", "기술", "대안", "선정 이유"]
for i, h in enumerate(tech_headers, 1):
    ws4.cell(row=4, column=i, value=h)
style_header_row(ws4, 4, 4)

techs = [
    ("Frontend", "React + TypeScript", "Vue.js, Angular", "차트 라이브러리 풍부, 컴포넌트 재사용성"),
    ("차트 라이브러리", "Apache ECharts", "Recharts, Chart.js", "실시간 스트리밍 + 대용량 데이터 렌더링 성능"),
    ("UI 프레임워크", "Ant Design", "MUI, Chakra UI", "테이블/모달/폼 등 관리 화면에 적합"),
    ("Backend", "Node.js (Fastify)", "Express, Spring Boot", "REST API, 비동기 처리에 강점"),
    ("실시간 통신", "WebSocket (Socket.io)", "SSE", "Data Monitoring 실시간 차트 지원"),
    ("Database", "MySQL", "PostgreSQL", "PDF 기능정의서에서 MySQL로 명시"),
    ("ORM", "Prisma", "TypeORM, Sequelize", "타입 안전한 DB 접근, 마이그레이션 관리"),
    ("인증", "JWT + bcrypt", "Session 기반", "Stateless REST API에 적합"),
    ("수학/신호처리", "math.js + 서버사이드", "-", "스무딩 필터, 적분, 선량 계산"),
]

for i, t in enumerate(techs):
    r = 5 + i
    for c, v in enumerate(t, 1):
        ws4.cell(row=r, column=c, value=v)
    style_data_row(ws4, r, 4)

# API 설계
api_start = 5 + len(techs) + 2
ws4.cell(row=api_start, column=1, value="API 엔드포인트 설계")
ws4.merge_cells(f"A{api_start}:D{api_start}")
style_category_row(ws4, api_start, 4)

api_headers = ["메서드", "엔드포인트", "설명", "비고"]
for i, h in enumerate(api_headers, 1):
    ws4.cell(row=api_start + 1, column=i, value=h)
style_header_row(ws4, api_start + 1, 4)

apis = [
    ("인증", None, None, None),
    ("POST", "/api/auth/login", "로그인", "JWT 토큰 반환"),
    ("POST", "/api/auth/logout", "로그아웃", ""),
    ("사용자", None, None, None),
    ("GET", "/api/users", "사용자 목록 조회", ""),
    ("POST", "/api/users", "사용자 추가", ""),
    ("PUT", "/api/users/:id", "사용자 수정", ""),
    ("DELETE", "/api/users/:id", "사용자 삭제", ""),
    ("PUT", "/api/users/:id/password", "비밀번호 변경", ""),
    ("Dosimeter 디바이스", None, None, None),
    ("GET", "/api/devices", "디바이스 목록", "페이지네이션"),
    ("POST", "/api/devices", "디바이스 추가", ""),
    ("GET", "/api/devices/:id", "디바이스 상세", ""),
    ("PUT", "/api/devices/:id", "디바이스 수정", ""),
    ("DELETE", "/api/devices/:id", "디바이스 삭제", ""),
    ("Gateway", None, None, None),
    ("GET", "/api/gateways", "Gateway 목록", ""),
    ("POST", "/api/gateways", "Gateway 추가", ""),
    ("GET", "/api/gateways/:id", "Gateway 상세", ""),
    ("PUT", "/api/gateways/:id/settings", "Gateway 설정 변경", "네트워크/BLE/LED 설정"),
    ("POST", "/api/gateways/:id/firmware/device", "Device FW 업로드", "multipart/form-data"),
    ("POST", "/api/gateways/:id/firmware/ble", "BLE FW 업로드", "multipart/form-data"),
    ("데이터 수집/조회", None, None, None),
    ("POST", "/api/data/ingest", "Gateway→서버 센서 데이터 전송", "Gateway가 호출"),
    ("GET", "/api/sensor-data", "센서 데이터 조회", "?deviceId=&startDate=&endDate=&startTime=&endTime="),
    ("실시간 모니터링", None, None, None),
    ("WS", "/ws/monitoring/:deviceId", "WebSocket 실시간 데이터", ""),
    ("Calibration", None, None, None),
    ("POST", "/api/calibrations/calculate", "스무딩 + 적분 계산", "필터, 윈도우, 베이스라인 파라미터"),
    ("POST", "/api/calibrations", "CF Factor 저장", ""),
    ("GET", "/api/calibrations", "Calibration 목록", ""),
    ("GET", "/api/calibrations/:id", "Calibration 상세", ""),
    ("PUT", "/api/calibrations/:id", "Calibration 수정", ""),
    ("DELETE", "/api/calibrations/:id", "Calibration 삭제", ""),
    ("Data Analysis", None, None, None),
    ("POST", "/api/analysis/calculate", "선량 계산", "방사선 종류, 장기, 범위 파라미터"),
    ("GET", "/api/analysis/export", "CSV 내보내기", "?id="),
    ("참조 데이터", None, None, None),
    ("GET", "/api/reference/radiation-factors", "방사선 가중치 목록", ""),
    ("GET", "/api/reference/tissue-factors", "조직 가중치 목록", ""),
]

r = api_start + 2
for a in apis:
    if a[1] is None:  # category
        ws4.cell(row=r, column=1, value=a[0])
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        style_category_row(ws4, r, 4)
    else:
        for c, v in enumerate(a, 1):
            ws4.cell(row=r, column=c, value=v)
        style_data_row(ws4, r, 4)
    r += 1

# 계산 로직
calc_start = r + 1
ws4.cell(row=calc_start, column=1, value="핵심 계산 로직")
ws4.merge_cells(f"A{calc_start}:D{calc_start}")
style_category_row(ws4, calc_start, 4)

calc_headers = ["단계", "수식", "단위", "설명"]
for i, h in enumerate(calc_headers, 1):
    ws4.cell(row=calc_start + 1, column=i, value=h)
style_header_row(ws4, calc_start + 1, 4)

calcs = [
    ("1. 누적선량", "Cumulative Dose = ∫(Start→End) [smoothed_voltage - baseline] dt", "V·s (mV·s)", "스무딩된 전압에서 베이스라인을 뺀 값의 시간 적분"),
    ("2. CF Factor", "CF Factor = Cumulative Dose / Delivered Dose", "mV·s / cGy", "누적선량을 전달선량으로 나눈 교정 계수"),
    ("3. 흡수선량", "Absorbed Dose = Cumulative Dose / CF Factor", "Gy (cGy)", "실제 흡수된 방사선량"),
    ("4. 등가선량", "Equivalent Dose = Absorbed Dose × wR", "Sv", "방사선 종류에 따른 가중치 적용"),
    ("5. 유효선량", "Effective Dose = Equivalent Dose × wT", "Sv", "장기별 가중치 적용한 최종 선량"),
]

for i, cl in enumerate(calcs):
    r = calc_start + 2 + i
    for c, v in enumerate(cl, 1):
        ws4.cell(row=r, column=c, value=v)
    style_data_row(ws4, r, 4)
    ws4.row_dimensions[r].height = 30

ws4.column_dimensions["A"].width = 20
ws4.column_dimensions["B"].width = 45
ws4.column_dimensions["C"].width = 25
ws4.column_dimensions["D"].width = 40

# ═══════════════════════════════════════════
# Sheet 5: 개발 로드맵
# ═══════════════════════════════════════════
ws5 = wb.create_sheet("5. 개발 로드맵")
ws5.sheet_properties.tabColor = "7030A0"

ws5.merge_cells("A1:E1")
ws5["A1"] = "Dosimetry 웹 애플리케이션 - 개발 로드맵"
ws5["A1"].font = TITLE_FONT
ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 40

roadmap_headers = ["Phase", "범위", "주요 작업", "산출물", "의존성"]
for i, h in enumerate(roadmap_headers, 1):
    ws5.cell(row=3, column=i, value=h)
style_header_row(ws5, 3, 5)

phases = [
    ("Phase 1\n기반 구축", "인증 + 디바이스 CRUD\n+ DB 구축",
     "• 프로젝트 초기 세팅 (React + Node.js + MySQL)\n• DB 스키마 생성 (Prisma 마이그레이션)\n• 로그인/인증 API (JWT)\n• Device CRUD API + UI (Dosimeter/Gateway)\n• 레이아웃 (Sidebar + Header)",
     "• Login 화면\n• Device 목록/상세 화면\n• DB 스키마\n• REST API 기본 구조",
     "없음"),
    ("Phase 2\n실시간 모니터링", "데이터 수집 +\n실시간 모니터링",
     "• Gateway 데이터 수신 API (POST /api/data/ingest)\n• sensor_data 테이블 적재\n• WebSocket 서버 구축\n• Data Monitoring UI (실시간 차트)\n• 디바이스 선택 + Start/Stop",
     "• Data Monitoring 화면\n• 실시간 Voltage 차트\n• 데이터 수집 파이프라인",
     "Phase 1"),
    ("Phase 3\nCalibration", "Calibration\n전체 플로우",
     "• 센서 데이터 조회 API\n• 스무딩 필터 6종 구현 (서버사이드)\n• 적분 계산 (Cumulative Dose)\n• CF Factor 산출 로직\n• Calibration UI (차트 + 필터 툴바 + Footer)\n• CF Factor 저장 API",
     "• Calibration 화면\n• 스무딩 필터 엔진\n• CF Factor 저장",
     "Phase 2"),
    ("Phase 4\n분석 기능", "Manage Calibration +\nData Analysis",
     "• Manage Calibration UI (CF Factor 목록/상세)\n• 방사선/조직 가중치 참조 테이블\n• 4종 선량 계산 로직\n• Data Analysis UI\n• CSV Export 기능",
     "• Manage Calibration 화면\n• Data Analysis 화면\n• CSV 다운로드",
     "Phase 3"),
    ("Phase 5\n완성 및 마무리", "사용자 관리 +\n워크스페이스 + 폴리싱",
     "• 사용자 CRUD API + UI\n• 워크스페이스 관리\n• 권한 기반 접근 제어\n• UI/UX 폴리싱\n• 에러 처리 + 로딩 상태\n• 배포 설정",
     "• Setting 모달\n• 사용자 관리 화면\n• 최종 배포본",
     "Phase 4"),
]

for i, p in enumerate(phases):
    r = 4 + i
    for c, v in enumerate(p, 1):
        ws5.cell(row=r, column=c, value=v)
    style_data_row(ws5, r, 5)
    ws5.row_dimensions[r].height = 120

ws5.column_dimensions["A"].width = 16
ws5.column_dimensions["B"].width = 20
ws5.column_dimensions["C"].width = 45
ws5.column_dimensions["D"].width = 30
ws5.column_dimensions["E"].width = 14

# ═══════════════════════════════════════════
# Sheet 6: 라이선스 및 비용 분석
# ═══════════════════════════════════════════
ws6 = wb.create_sheet("6. 라이선스 및 비용")
ws6.sheet_properties.tabColor = "00B050"

ws6.merge_cells("A1:G1")
ws6["A1"] = "Dosimetry 웹 애플리케이션 - 라이선스 및 비용 분석"
ws6["A1"].font = TITLE_FONT
ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 40

# ── 기술 스택 라이선스 ──
ws6.cell(row=3, column=1, value="기술 스택 라이선스 현황")
ws6.merge_cells("A3:G3")
style_category_row(ws6, 3, 7)

lic_headers = ["분류", "기술", "라이선스", "비용", "상업적 사용", "소스코드 공개 의무", "비고"]
for i, h in enumerate(lic_headers, 1):
    ws6.cell(row=4, column=i, value=h)
style_header_row(ws6, 4, 7)

FREE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NOTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

licenses = [
    ("Frontend", "React", "MIT", "무료", "O", "X", "Meta(Facebook) 개발, 제한 없이 사용 가능"),
    ("Frontend", "TypeScript", "Apache 2.0", "무료", "O", "X", "Microsoft 개발"),
    ("차트", "Apache ECharts", "Apache 2.0", "무료", "O", "X", "Apache Software Foundation 프로젝트"),
    ("UI", "Ant Design", "MIT", "무료", "O", "X", "Alibaba 개발"),
    ("Backend", "Node.js", "MIT", "무료", "O", "X", "OpenJS Foundation"),
    ("Backend", "Fastify", "MIT", "무료", "O", "X", "Express 대비 고성능"),
    ("실시간", "Socket.io", "MIT", "무료", "O", "X", "WebSocket 라이브러리"),
    ("DB", "MySQL Community Edition", "GPL v2", "무료", "O", "조건부 *", "* 서버 내부 사용 시 공개 의무 없음\n(배포하지 않는 한 해당 없음)"),
    ("ORM", "Prisma", "Apache 2.0", "무료", "O", "X", "Prisma Cloud(호스팅)만 유료,\nORM 자체는 완전 무료"),
    ("인증", "jsonwebtoken (JWT)", "MIT", "무료", "O", "X", "토큰 기반 인증"),
    ("인증", "bcrypt", "MIT", "무료", "O", "X", "비밀번호 해싱"),
    ("수학", "math.js", "Apache 2.0", "무료", "O", "X", "수학 연산 라이브러리"),
    ("빌드", "Vite", "MIT", "무료", "O", "X", "프론트엔드 빌드 도구"),
    ("기타", "ESLint", "MIT", "무료", "O", "X", "코드 품질 검사"),
    ("기타", "Prettier", "MIT", "무료", "O", "X", "코드 포맷터"),
]

for i, lic in enumerate(licenses):
    r = 5 + i
    for c, v in enumerate(lic, 1):
        ws6.cell(row=r, column=c, value=v)
    style_data_row(ws6, r, 7)
    # 비용 컬럼 초록 배경
    ws6.cell(row=r, column=4).fill = FREE_FILL
    ws6.cell(row=r, column=4).alignment = CENTER_ALIGN
    # 상업적 사용 O 표시
    ws6.cell(row=r, column=5).alignment = CENTER_ALIGN
    ws6.cell(row=r, column=6).alignment = CENTER_ALIGN
    ws6.row_dimensions[r].height = 30

# ── 라이선스 유형 설명 ──
lic_type_start = 5 + len(licenses) + 2
ws6.cell(row=lic_type_start, column=1, value="라이선스 유형별 설명")
ws6.merge_cells(f"A{lic_type_start}:G{lic_type_start}")
style_category_row(ws6, lic_type_start, 7)

lic_type_headers = ["라이선스", "전체 이름", "핵심 조건", "상업적 사용", "소스 공개 의무", "특징", ""]
for i, h in enumerate(lic_type_headers, 1):
    ws6.cell(row=lic_type_start + 1, column=i, value=h)
style_header_row(ws6, lic_type_start + 1, 7)

lic_types = [
    ("MIT", "MIT License", "저작권 표시만 유지", "가능", "없음", "가장 자유로운 오픈소스 라이선스. 수정, 배포, 상업적 사용 모두 자유.", ""),
    ("Apache 2.0", "Apache License 2.0", "저작권 표시 + 변경사항 명시", "가능", "없음", "MIT와 유사하나 특허권 보호 조항 포함. 기업 친화적.", ""),
    ("GPL v2", "GNU General Public License v2", "동일 라이선스 적용\n(배포 시)", "가능", "배포 시에만", "소프트웨어를 외부에 배포할 때만 소스 공개 의무 발생.\n서버 내부 사용(SaaS)은 해당 없음.", ""),
]

for i, lt in enumerate(lic_types):
    r = lic_type_start + 2 + i
    for c, v in enumerate(lt, 1):
        ws6.cell(row=r, column=c, value=v)
    style_data_row(ws6, r, 7)
    ws6.row_dimensions[r].height = 45

# ── 인프라/운영 비용 ──
infra_start = lic_type_start + 2 + len(lic_types) + 2
ws6.cell(row=infra_start, column=1, value="인프라/운영 비용 (기술 스택 외)")
ws6.merge_cells(f"A{infra_start}:G{infra_start}")
style_category_row(ws6, infra_start, 7)

infra_headers = ["항목", "무료 옵션", "유료 시 비용", "현재 시스템 적용", "비고", "", ""]
for i, h in enumerate(infra_headers, 1):
    ws6.cell(row=infra_start + 1, column=i, value=h)
style_header_row(ws6, infra_start + 1, 7)

infra = [
    ("서버 호스팅", "사내 서버 (On-Premise)", "AWS EC2: 월 $10~50\nGCP: 월 $10~50", "사내 서버 (192.168.0.200)", "내부 네트워크 운영 시 추가 비용 없음", "", ""),
    ("도메인", "내부 IP 사용 시 불필요", "외부 도메인: 연 1~2만원", "내부 IP 사용 중", "외부 공개 시에만 필요", "", ""),
    ("SSL 인증서", "Let's Encrypt (무료)", "유료 SSL: 연 5~20만원", "내부망이면 불필요", "외부 공개 시 Let's Encrypt로 무료 해결 가능", "", ""),
    ("DB 호스팅", "사내 서버에 MySQL 직접 설치", "AWS RDS: 월 $15~100\nPlanetScale: 무료 티어 있음", "사내 서버 직접 설치", "MySQL Community로 충분", "", ""),
    ("이메일 서비스", "불필요 (현재 기능에 없음)", "필요 시: SendGrid 무료 티어\n(일 100건)", "해당 없음", "향후 알림 기능 추가 시 검토", "", ""),
]

for i, inf in enumerate(infra):
    r = infra_start + 2 + i
    for c, v in enumerate(inf, 1):
        ws6.cell(row=r, column=c, value=v)
    style_data_row(ws6, r, 7)
    ws6.row_dimensions[r].height = 40

# ── 결론 ──
conclusion_start = infra_start + 2 + len(infra) + 2
ws6.cell(row=conclusion_start, column=1, value="결론")
ws6.merge_cells(f"A{conclusion_start}:G{conclusion_start}")
style_category_row(ws6, conclusion_start, 7)

CONCLUSION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
CONCLUSION_FONT = Font(name="Arial", size=11, bold=True, color="006100")

conclusions = [
    "1. 제안된 모든 기술 스택은 MIT 또는 Apache 2.0 라이선스로, 상업적 사용 포함 완전 무료입니다.",
    "2. MySQL은 GPL v2이나, 서버 내부 사용(외부 배포하지 않음) 시 소스 공개 의무가 없습니다.",
    "3. 현재 시스템이 사내 서버(192.168.0.200)에서 운영되므로, 별도의 클라우드/인프라 비용이 발생하지 않습니다.",
    "4. 라이선스 비용: 0원  |  인프라 비용(사내 서버): 0원  |  총 추가 비용: 0원",
    "5. 유일한 주의사항: MySQL을 수정하여 외부에 재배포할 경우에만 GPL 조건이 적용됩니다. (일반적 사용에는 해당 없음)",
]

for i, con in enumerate(conclusions):
    r = conclusion_start + 1 + i
    ws6.cell(row=r, column=1, value=con)
    ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws6.cell(row=r, column=1).font = CONCLUSION_FONT
    ws6.cell(row=r, column=1).fill = CONCLUSION_FILL
    ws6.cell(row=r, column=1).alignment = WRAP_ALIGN
    ws6.cell(row=r, column=1).border = THIN_BORDER
    ws6.row_dimensions[r].height = 25

ws6.column_dimensions["A"].width = 18
ws6.column_dimensions["B"].width = 25
ws6.column_dimensions["C"].width = 28
ws6.column_dimensions["D"].width = 22
ws6.column_dimensions["E"].width = 25
ws6.column_dimensions["F"].width = 22
ws6.column_dimensions["G"].width = 20

# ── 저장 ──
output_path = "/Users/ws_eric/Documents/rumi/Dosimetry_개발플랜.xlsx"
wb.save(output_path)
print(f"Excel saved: {output_path}")
